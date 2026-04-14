"""eval/script.py - Comprehensive benchmark for all 4 agents.

Covers all 35 tools (Graph: 11, Salesforce: 5, SmartSales: 19) across 4 agent
modes. After collecting responses an LLM evaluator scores each one (1–5) by
comparing it to a human-written expected answer.

Output: benchmark_results.xlsx with one sheet per agent + a Summary sheet.
Each run appends new rows (identified by run_id) so results accumulate.

Usage (from project root):
    python eval/script.py
"""

import asyncio
import configparser
import json
import os
import socket
import subprocess
import sys
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime
import webbrowser
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import parse_qs, urlparse

import httpx
import msal
import openpyxl
from dotenv import load_dotenv
from openai import AsyncAzureOpenAI
from openpyxl.styles import Font, PatternFill

from agent_framework import MCPStreamableHTTPTool
from agents.graph_agent import create_graph_agent
from agents.orchestrator_agent import create_orchestrator_agent
from agents.routing_trace import get_trace, start_trace
from agents.salesforce_agent import create_salesforce_agent
from agents.smartsales_agent import create_smartsales_agent

load_dotenv()


# ── Prompt dataclass ──────────────────────────────────────────────────────────

@dataclass
class Prompt:
    text: str
    category: str       # e.g. "email", "locations", "cross-system"
    difficulty: str     # "simple" | "medium" | "hard"
    expected_answer: str = ""   # human-written description of a correct response
    tags: list[str] = field(default_factory=list)


# ── Graph prompts - 11 tools ──────────────────────────────────────────────────

GRAPH_PROMPTS: list[Prompt] = [
    # whoami
    Prompt(
        text="Who am I in Microsoft 365?",
        category="identity",
        difficulty="simple",
        expected_answer=(
            "The response contains the authenticated user's display name and "
            "email address from Microsoft 365."
        ),
        tags=["whoami"],
    ),
    # list_email
    Prompt(
        text="Show me my 5 most recent emails.",
        category="email",
        difficulty="simple",
        expected_answer=(
            "A list of the 5 most recent inbox emails, each with at minimum the "
            "subject line, sender name or email address, and received date/time."
        ),
        tags=["list_email"],
    ),
    # search_email - by subject keyword
    Prompt(
        text="Search for emails that have the word 'meeting' in the subject.",
        category="email",
        difficulty="medium",
        expected_answer=(
            "A list of emails whose subject contains 'meeting', showing subject, "
            "sender, and date. If no results are found, the response clearly states so."
        ),
        tags=["search_email"],
    ),
    # search_email - by date range
    Prompt(
        text="Have I received any emails in the last 7 days? List sender and subject.",
        category="email",
        difficulty="medium",
        expected_answer=(
            "A list of emails received in the past 7 days, each showing the sender "
            "and subject. If none, the response clearly states the inbox was empty for that period."
        ),
        tags=["search_email"],
    ),
    # read_email - chained: list_email → read_email
    Prompt(
        text="What does my most recent email say? Give me the full body.",
        category="email",
        difficulty="hard",
        expected_answer=(
            "The full body text of the most recent email, preceded by its subject "
            "and sender. The body is not truncated or summarised."
        ),
        tags=["list_email", "read_email"],
    ),
    # findpeople
    Prompt(
        text="Find the email address of Dorian.",
        category="people",
        difficulty="simple",
        expected_answer=(
            "The email address(es) associated with a person named Dorian found in "
            "the Microsoft 365 directory, along with their display name."
        ),
        tags=["findpeople"],
    ),
    # list_calendar
    Prompt(
        text="What are my upcoming calendar events this week?",
        category="calendar",
        difficulty="simple",
        expected_answer=(
            "A list of calendar events for the current week, each with event title, "
            "start date and time, and optionally end time or location."
        ),
        tags=["list_calendar"],
    ),
    # search_calendar
    Prompt(
        text="Search my calendar for any events or meetings in the next 14 days.",
        category="calendar",
        difficulty="medium",
        expected_answer=(
            "A list of calendar events occurring in the next 14 days, each showing "
            "the event title and start date/time"
        ),
        tags=["search_calendar"],
    ),
    # list_contacts
    Prompt(
        text="Show me my Microsoft 365 contacts.",
        category="contacts",
        difficulty="simple",
        expected_answer=(
            "A list of the user's Microsoft 365 contacts, each with display name "
            "and email address."
        ),
        tags=["list_contacts"],
    ),
    # search_files
    Prompt(
        text="Find any Excel or PDF files in my OneDrive.",
        category="files",
        difficulty="medium",
        expected_answer=(
            "A list of Excel (.xlsx/.xls) or PDF files found in OneDrive, each "
            "with the file name and its ID."
        ),
        tags=["search_files"],
    ),
    # search_files → read_file (chained)
    Prompt(
        text="Search OneDrive for a file called 'report' and read its content.",
        category="files",
        difficulty="hard",
        expected_answer=(
            "The text content of a file whose name contains 'report', preceded by "
            "the file name. If no such file is found, the response clearly states so."
        ),
        tags=["search_files", "read_file"],
    ),
    # search_email - sender filter
    Prompt(
        text="Show me emails I received in the last 14 days.",
        category="email",
        difficulty="medium",
        expected_answer=(
            "A list of emails received in the past 14 days, each showing at minimum "
            "subject, sender, and received date, , mind the format: year-month-day"
        ),
        tags=["search_email"],
    ),
    # search_email - subject filter
    Prompt(
        text="Find emails with 'invoice' in the subject.",
        category="email",
        difficulty="medium",
        expected_answer=(
            "A list of emails where the subject contains 'invoice', showing sender, "
            "subject, and date. If none found, the response states so clearly."
        ),
        tags=["search_email"],
    ),
    # search_email → read_email (chained, hard)
    Prompt(
        text="Find the most recent email with 'meeting' in the subject and read its full body.",
        category="email",
        difficulty="hard",
        expected_answer=(
            "The full body of the most recent email whose subject contains 'meeting', "
            "preceded by the sender and subject line."
        ),
        tags=["search_email", "read_email"],
    ),
    # findpeople - second variant
    Prompt(
        text="Look up the email address of Arne in the organization.",
        category="people",
        difficulty="simple",
        expected_answer=(
            "The email address and display name of one or more people named Arne"
            "found in the Microsoft 365 directory."
        ),
        tags=["findpeople"],
    ),
    # search_files - Word documents
    Prompt(
        text="Find all Word documents in my OneDrive.",
        category="files",
        difficulty="simple",
        expected_answer=(
            "A list of Word (.docx or .doc) files found in OneDrive, each with the "
            "file name and its ID."
        ),
        tags=["search_files"],
    ),
    # search_files → read_file - medium chained
    Prompt(
        text="Find a document called 'agenda' in OneDrive and show its content.",
        category="files",
        difficulty="medium",
        expected_answer=(
            "The text content of a file named 'agenda' retrieved from OneDrive. "
        ),
        tags=["search_files", "read_file"],
    ),
    # search_files → read_multiple_files (chained, hard)
    Prompt(
        text="Search OneDrive for Excel files and read the content of the first two found.",
        category="files",
        difficulty="hard",
        expected_answer=(
            "The text content of the first two Excel files found in OneDrive, "
            "each clearly labelled with its filename."
        ),
        tags=["search_files", "read_multiple_files"],
    ),
    # search_files → read_multiple_files - broader search
    Prompt(
        text="Find any files related to 'nutella' in OneDrive and read all of them.",
        category="files",
        difficulty="hard",
        expected_answer=(
            "The text content of every file matching 'nutella' found in OneDrive, "
            "each labelled with its filename. If no files are found, states so clearly."
        ),
        tags=["search_files", "read_multiple_files"],
    ),
    # list_contacts - with phone numbers
    Prompt(
        text="List my Microsoft 365 contacts and include their phone numbers.",
        category="contacts",
        difficulty="medium",
        expected_answer=(
            "A list of Microsoft 365 contacts, each showing display name, email "
            "address, and phone number (if available)."
        ),
        tags=["list_contacts"],
    ),
    # list_calendar - today
    Prompt(
        text="How many calendar events do I have scheduled this week? List them all.",
        category="calendar",
        difficulty="medium",
        expected_answer=(
            "A count and a complete list of all calendar events in the current week, "
            "each with event title and start date/time."
        ),
        tags=["list_calendar"],
    ),
    # search_calendar - next 30 days
    Prompt(
        text="Search my calendar for events scheduled in the next 30 days.",
        category="calendar",
        difficulty="medium",
        expected_answer=(
            "A list of calendar events occurring in the next 30 days, each showing "
            "event title, start date/time, and optionally location. Today is the 14th of april 2026"
        ),
        tags=["search_calendar"],
    ),
]


# ── Salesforce prompts - 5 tools ──────────────────────────────────────────────

SALESFORCE_PROMPTS: list[Prompt] = [
    # find_accounts - basic
    Prompt(
        text="List 5 Salesforce accounts.",
        category="accounts",
        difficulty="simple",
        expected_answer=(
            "A list of exactly 5 Salesforce account records, each with at minimum "
            "the Id and Name fields."
        ),
        tags=["find_accounts"],
    ),
    # find_accounts - extra_fields + filter
    Prompt(
        text="Find Salesforce accounts in Belgium, including their billing address.",
        category="accounts",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce accounts where BillingCountry is Belgium, each "
            "showing Name and billing address fields (BillingStreet, BillingCity, BillingCountry)."
        ),
        tags=["find_accounts"],
    ),
    # find_contacts - basic
    Prompt(
        text="Show me 5 Salesforce contacts with their email addresses.",
        category="contacts",
        difficulty="simple",
        expected_answer=(
            "A list of 5 Salesforce contact records, each with Name and Email fields."
        ),
        tags=["find_contacts"],
    ),
    # find_contacts - filter
    Prompt(
        text="Find Salesforce contacts in the Sales department.",
        category="contacts",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce contacts where Department equals 'Sales', each "
            "showing Name and Department. If none found, the response states so clearly."
        ),
        tags=["find_contacts"],
    ),
    # find_leads - basic
    Prompt(
        text="Show me 5 leads in Salesforce.",
        category="leads",
        difficulty="simple",
        expected_answer=(
            "A list of 5 Salesforce lead records, each with at minimum Name and Company fields."
        ),
        tags=["find_leads"],
    ),
    # find_leads - industry filter
    Prompt(
        text="Find Salesforce leads from the Technology or Software industry.",
        category="leads",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce leads where Industry is Technology or Software, "
            "each showing Name and Industry."
        ),
        tags=["find_leads"],
    ),
    # get_opportunities - basic
    Prompt(
        text="List 5 open opportunities in Salesforce.",
        category="opportunities",
        difficulty="simple",
        expected_answer=(
            "A list of 5 open Salesforce opportunities, each with Name and StageName fields."
        ),
        tags=["get_opportunities"],
    ),
    # get_opportunities - amount filter + extra field
    Prompt(
        text="Show me Salesforce opportunities with an amount greater than 10,000. Include probability.",
        category="opportunities",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce opportunities where Amount > 10000, each showing "
            "Name, Amount, and Probability."
        ),
        tags=["get_opportunities"],
    ),
    # get_cases - open
    Prompt(
        text="List open support cases in Salesforce.",
        category="cases",
        difficulty="simple",
        expected_answer=(
            "A list of open Salesforce cases (Status != Closed), each with "
            "CaseNumber and Status fields."
        ),
        tags=["get_cases"],
    ),
    # get_cases - closed + extra fields
    Prompt(
        text="Show me 5 closed Salesforce cases including their close date and description.",
        category="cases",
        difficulty="medium",
        expected_answer=(
            "A list of 5 closed Salesforce cases, each with CaseNumber, Status "
            "(Closed), ClosedDate, and Description fields."
        ),
        tags=["get_cases"],
    ),
    # find_accounts - name contains filter
    Prompt(
        text="Find Salesforce accounts that have 'Group' in their name.",
        category="accounts",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce accounts where Name contains 'Group', each "
            "showing Id and Name."
        ),
        tags=["find_accounts"],
    ),
    # find_accounts - extra fields
    Prompt(
        text="Show me Salesforce accounts with their phone numbers.",
        category="accounts",
        difficulty="simple",
        expected_answer=(
            "A list of Salesforce accounts each showing Id, Name, and Phone. "
            "Accounts without a phone number may show an empty value."
        ),
        tags=["find_accounts"],
    ),
    # find_contacts - title filter
    Prompt(
        text="Find Salesforce contacts with 'Manager' in their title.",
        category="contacts",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce contacts where Title contains 'Manager', each "
            "showing Name, Title, and Email."
        ),
        tags=["find_contacts"],
    ),
    # find_contacts - recently created
    Prompt(
        text="Show me the most recently created Salesforce contacts.",
        category="contacts",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce contacts each showing Name, Email"
        ),
        tags=["find_contacts"],
    ),
    # find_leads - converted
    Prompt(
        text="Find Salesforce leads that have been converted.",
        category="leads",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce leads where IsConverted is true, each showing "
            "Name, Company, and Status."
        ),
        tags=["find_leads"],
    ),
    # find_leads - status filter
    Prompt(
        text="Show me Salesforce leads with the status 'New'.",
        category="leads",
        difficulty="simple",
        expected_answer=(
            "A list of Salesforce leads where Status is 'new', "
            "each showing Name, Company, and Status."
        ),
        tags=["find_leads"],
    ),
    # get_opportunities - stage filter
    Prompt(
        text="Show me Salesforce opportunities in the 'Qualification' stage.",
        category="opportunities",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce opportunities where StageName is "
            "'Qualification', each showing Name, Amount, and StageName."
        ),
        tags=["get_opportunities"],
    ),
    # get_opportunities - sorted by amount
    Prompt(
        text="List Salesforce opportunities sorted by amount, highest first.",
        category="opportunities",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce opportunities ordered by Amount descending, "
            "each showing Name, Amount, and StageName."
        ),
        tags=["get_opportunities"],
    ),
    # get_cases - high priority
    Prompt(
        text="Find high-priority open cases in Salesforce.",
        category="cases",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce cases where Status is not Closed and Priority "
            "is High, each showing CaseNumber, Subject, and Priority."
        ),
        tags=["get_cases"],
    ),
    # get_cases - with account names
    Prompt(
        text="Show me Salesforce cases with their related account names.",
        category="cases",
        difficulty="medium",
        expected_answer=(
            "A list of Salesforce cases each showing CaseNumber, Status, Subject, "
            "and the Name of the related Account."
        ),
        tags=["get_cases"],
    ),
    # find_accounts → find_contacts (chained)
    Prompt(
        text="Find Salesforce accounts in France and list the contacts linked to those accounts.",
        category="accounts",
        difficulty="hard",
        expected_answer=(
            "First a list of Salesforce accounts where BillingCountry is France, "
            "then the contacts associated with each of those accounts."
        ),
        tags=["find_accounts", "find_contacts"],
    ),
    # get_opportunities → find_accounts (chained)
    Prompt(
        text="Show me the 3 largest open Salesforce opportunities and the billing country of each related account.",
        category="opportunities",
        difficulty="hard",
        expected_answer=(
            "The 3 open Salesforce opportunities with the highest Amount, each "
            "showing Name, Amount, and the BillingCountry of the related Account."
        ),
        tags=["get_opportunities", "find_accounts"],
    ),
]


# ── SmartSales prompts - all 19 tools ────────────────────────────────────────

SMARTSALES_PROMPTS: list[Prompt] = [
    # ── LOCATIONS ─────────────────────────────────────────────────────────────
    # list_locations - basic
    Prompt(
        text="List all SmartSales locations.",
        category="locations",
        difficulty="simple",
        expected_answer=(
            "A JSON array of SmartSales location objects. Each object contains at "
            "minimum the uid and name fields. The response also reports the total count."
        ),
        tags=["list_locations"],
    ),
    # list_locations - city filter
    Prompt(
        text="Find SmartSales locations in Brussels.",
        category="locations",
        difficulty="medium",
        expected_answer=(
            "A JSON array of SmartSales locations where city equals Brussels. Each "
            "object has at minimum uid and name. If none found, states so clearly."
        ),
        tags=["list_locations"],
    ),
    # list_locations - sort + projection
    Prompt(
        text="List SmartSales locations in Belgium sorted by name, using full projection.",
        category="locations",
        difficulty="medium",
        expected_answer=(
            "A JSON array of SmartSales locations in Belgium (country eq Belgium), "
            "sorted alphabetically by name, with full projection fields for each entry."
        ),
        tags=["list_locations"],
    ),
    # list_locations → get_location (chained)
    Prompt(
        text="List all SmartSales locations, then retrieve the full details of the first result.",
        category="locations",
        difficulty="hard",
        expected_answer=(
            "First the list of locations, then the complete detail object of the "
            "first location retrieved by its uid, containing all available fields."
        ),
        tags=["list_locations", "get_location"],
    ),
    # list_displayable_fields
    Prompt(
        text="What fields are available to display in the SmartSales location list?",
        category="locations",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors indicating which fields can be "
            "shown in the SmartSales location list view (e.g. name, city, country, uid)."
        ),
        tags=["list_displayable_fields"],
    ),
    # list_queryable_fields
    Prompt(
        text="What fields can I use to filter SmartSales locations?",
        category="locations",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields support "
            "query/filter operators for SmartSales locations (e.g. city, country, name)."
        ),
        tags=["list_queryable_fields"],
    ),
    # list_sortable_fields
    Prompt(
        text="What fields can I use to sort SmartSales locations?",
        category="locations",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields can be used "
            "as sort keys for SmartSales locations."
        ),
        tags=["list_sortable_fields"],
    ),
    # ── CATALOG ───────────────────────────────────────────────────────────────
    # list_catalog_items - basic
    Prompt(
        text="List SmartSales catalog items.",
        category="catalog",
        difficulty="simple",
        expected_answer=(
            "A JSON array of SmartSales catalog item objects. Each object contains "
            "at minimum the uid field."
        ),
        tags=["list_catalog_items"],
    ),
    # list_catalog_items - sorted + projection
    Prompt(
        text="List SmartSales catalog items with simple projection, sorted by name.",
        category="catalog",
        difficulty="medium",
        expected_answer=(
            "A JSON array of catalog items with simple projection fields, returned "
            "in alphabetical order by name."
        ),
        tags=["list_catalog_items"],
    ),
    # list_catalog_items → get_catalog_item (chained)
    Prompt(
        text="List SmartSales catalog items, then retrieve the full details of the first one.",
        category="catalog",
        difficulty="hard",
        expected_answer=(
            "First the list of catalog items, then the full detail object of the "
            "first item retrieved by its uid, containing all available fields."
        ),
        tags=["list_catalog_items", "get_catalog_item"],
    ),
    # list_catalog_displayable_fields
    Prompt(
        text="What fields can be displayed for SmartSales catalog items?",
        category="catalog",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors indicating which fields can be "
            "shown in the SmartSales catalog item list view."
        ),
        tags=["list_catalog_displayable_fields"],
    ),
    # list_catalog_queryable_fields
    Prompt(
        text="What fields can I filter SmartSales catalog items by?",
        category="catalog",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields support "
            "query/filter operators for SmartSales catalog items."
        ),
        tags=["list_catalog_queryable_fields"],
    ),
    # list_catalog_sortable_fields
    Prompt(
        text="What fields can I sort SmartSales catalog items by?",
        category="catalog",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields can be used "
            "as sort keys for SmartSales catalog items."
        ),
        tags=["list_catalog_sortable_fields"],
    ),
    # ── ORDERS ────────────────────────────────────────────────────────────────
    # list_orders - basic
    Prompt(
        text="List recent SmartSales orders.",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A JSON array of SmartSales order objects. Each object contains at "
            "minimum the uid field."
        ),
        tags=["list_orders"],
    ),
    # list_orders - full projection
    Prompt(
        text="List SmartSales orders with full projection.",
        category="orders",
        difficulty="medium",
        expected_answer=(
            "A JSON array of SmartSales orders with the full projection applied, "
            "meaning each order object contains all available fields."
        ),
        tags=["list_orders"],
    ),
    # list_orders → get_order (chained)
    Prompt(
        text="List SmartSales orders, then retrieve the full details of the first order.",
        category="orders",
        difficulty="hard",
        expected_answer=(
            "First the list of orders, then the complete detail object of the first "
            "order retrieved by its uid, containing all available fields."
        ),
        tags=["list_orders", "get_order"],
    ),
    # get_order_configuration
    Prompt(
        text="What is the SmartSales order configuration?",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A JSON object describing the global SmartSales order configuration, "
            "including form sections, configurable fields, and order settings."
        ),
        tags=["get_order_configuration"],
    ),
    # list_approbation_statuses
    Prompt(
        text="List all SmartSales order approbation statuses.",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A JSON array of SmartSales order approbation (approval) status objects, "
            "each with at minimum a uid field."
        ),
        tags=["list_approbation_statuses"],
    ),
    # list_approbation_statuses → get_approbation_status (chained)
    Prompt(
        text="List SmartSales approbation statuses, then get the full details of the first one.",
        category="orders",
        difficulty="hard",
        expected_answer=(
            "First the list of approbation statuses, then the full detail object of "
            "the first status retrieved by its uid."
        ),
        tags=["list_approbation_statuses", "get_approbation_status"],
    ),
    # list_order_displayable_fields
    Prompt(
        text="What fields can be displayed for SmartSales orders?",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors indicating which fields can be "
            "shown in the SmartSales order list view."
        ),
        tags=["list_order_displayable_fields"],
    ),
    # list_order_queryable_fields
    Prompt(
        text="What fields can I filter SmartSales orders by?",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields support "
            "query/filter operators for SmartSales orders."
        ),
        tags=["list_order_queryable_fields"],
    ),
    # list_order_sortable_fields
    Prompt(
        text="What fields can I sort SmartSales orders by?",
        category="orders",
        difficulty="simple",
        expected_answer=(
            "A list or array of field descriptors showing which fields can be used "
            "as sort keys for SmartSales orders."
        ),
        tags=["list_order_sortable_fields"],
    ),
]


# ── Orchestrator prompts - routing + cross-system ─────────────────────────────

_ORCHESTRATOR_ROUTING: list[Prompt] = [
    Prompt(
        text="Who am I in Microsoft 365?",
        category="routing/graph",
        difficulty="simple",
        expected_answer=(
            "The authenticated user's display name and email address from Microsoft "
            "365, routed correctly via the Graph agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    Prompt(
        text="List 5 Salesforce accounts.",
        category="routing/salesforce",
        difficulty="simple",
        expected_answer=(
            "A list of 5 Salesforce account records with Id and Name, routed "
            "correctly via the Salesforce agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    Prompt(
        text="List SmartSales locations.",
        category="routing/smartsales",
        difficulty="simple",
        expected_answer=(
            "A JSON array of SmartSales location objects with uid and name fields, "
            "routed correctly via the SmartSales agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    # routing/graph - calendar
    Prompt(
        text="What calendar events do I have next week?",
        category="routing/graph",
        difficulty="medium",
        expected_answer=(
            "A list of calendar events scheduled for next week, each with title "
            "and start date/time, retrieved via the Graph agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    # routing/graph - email search
    Prompt(
        text="Search my emails for anything about budget.",
        category="routing/graph",
        difficulty="medium",
        expected_answer=(
            "A list of emails where the subject or body contains 'budget', "
            "showing sender, subject, and date, retrieved via the Graph agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    # routing/salesforce - opportunities
    Prompt(
        text="Find all open Salesforce opportunities with an amount above 10,000.",
        category="routing/salesforce",
        difficulty="medium",
        expected_answer=(
            "A list of open Salesforce opportunities where Amount > 10000, each "
            "showing Name, Amount, and StageName, routed via the Salesforce agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    # routing/smartsales - catalog
    Prompt(
        text="List SmartSales catalog items.",
        category="routing/smartsales",
        difficulty="simple",
        expected_answer=(
            "A list of SmartSales catalog items, each with at minimum the uid "
            "field, routed correctly via the SmartSales agent."
        ),
        tags=["orchestrator", "routing"],
    ),
    # routing/smartsales - orders
    Prompt(
        text="Show me recent SmartSales orders.",
        category="routing/smartsales",
        difficulty="medium",
        expected_answer=(
            "A list of recent SmartSales orders, each with at minimum the uid "
            "field, routed correctly via the SmartSales agent."
        ),
        tags=["orchestrator", "routing"],
    ),
]

_ORCHESTRATOR_CROSS: list[Prompt] = [
    Prompt(
        text="Find contacts named 'John' in both Microsoft 365 and Salesforce.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Results from both systems, clearly labelled. From Microsoft 365: any "
            "contacts named John with their email. From Salesforce: any contacts "
            "named John with their email and account."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    Prompt(
        text="Show me my calendar events for the next 7 days and check whether any organizers appear as contacts in Salesforce.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "A list of calendar events for the next 7 days (title, date, organizer "
            "email), followed by a note for each organizer whether they were found "
            "in Salesforce contacts."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    Prompt(
        text="List SmartSales locations in Brussels and check if there are any matching Salesforce accounts in the same city.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "SmartSales locations in Brussels (uid, name) labelled 'From SmartSales', "
            "followed by Salesforce accounts with BillingCity Brussels labelled "
            "'From Salesforce'."
        ),
        tags=["orchestrator", "multi-agent", "smartsales", "salesforce"],
    ),
    Prompt(
        text="What are my 3 most recent emails? Are any of those senders listed as contacts in Salesforce?",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "The 3 most recent emails (subject, sender), then a check for each "
            "sender in Salesforce contacts, clearly stating whether each was found."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    Prompt(
        text="Show me open Salesforce opportunities and SmartSales locations in Belgium. Give me a summary of both.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Two clearly labelled sections: open Salesforce opportunities (Name, "
            "StageName, Amount) and SmartSales locations in Belgium (uid, name), "
            "followed by a short summary of each."
        ),
        tags=["orchestrator", "multi-agent", "salesforce", "smartsales"],
    ),
    # graph + salesforce - email senders as leads
    Prompt(
        text="Find my 5 most recent emails and check if any of the senders are Salesforce leads.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "The 5 most recent emails showing sender and subject, followed by a "
            "note for each sender on whether they appear as a lead in Salesforce."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    # graph + smartsales - contacts vs locations
    Prompt(
        text="List my Microsoft 365 contacts and check if any of them are also SmartSales locations.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "My Microsoft 365 contacts with display name and email, with a note "
            "for each on whether a matching name appears as a SmartSales location."
        ),
        tags=["orchestrator", "multi-agent", "graph", "smartsales"],
    ),
    # salesforce + smartsales - accounts vs locations same city
    Prompt(
        text="Find Salesforce accounts in Brussels and compare with SmartSales locations in the same city.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Salesforce accounts with BillingCity Brussels and SmartSales locations "
            "filtered by city Brussels, shown in two clearly labelled sections."
        ),
        tags=["orchestrator", "multi-agent", "salesforce", "smartsales"],
    ),
    # graph + salesforce - calendar attendees as contacts
    Prompt(
        text="Show me my calendar events for the next 7 days and check if any attendees are Salesforce contacts.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Calendar events for the next 7 days with title and attendees, followed "
            "by a note for each unique attendee email on whether they are a Salesforce contact."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    # graph + salesforce + smartsales - all three
    Prompt(
        text="Who am I in Microsoft 365, what Salesforce accounts exist, and what SmartSales locations are in Belgium?",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Three clearly labelled sections: (1) Microsoft 365 identity with "
            "display name and email, (2) a list of Salesforce accounts, "
            "(3) SmartSales locations in Belgium."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce", "smartsales"],
    ),
    # salesforce + smartsales - orders vs opportunities
    Prompt(
        text="List recent SmartSales orders and check if the customer names match any Salesforce accounts.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "A list of recent SmartSales orders with customer information, followed "
            "by a check per customer whether a matching account name exists in Salesforce."
        ),
        tags=["orchestrator", "multi-agent", "salesforce", "smartsales"],
    ),
    # graph + salesforce - files + accounts
    Prompt(
        text="Search my OneDrive for files about contracts and check if the mentioned company names are Salesforce accounts.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Files found in OneDrive matching 'contract' with their content summary, "
            "followed by a check for each mentioned company name in Salesforce accounts."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce"],
    ),
    # graph + smartsales - email senders vs locations
    Prompt(
        text="Find emails I received this week and check if any senders are linked to a SmartSales location.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Emails received this week showing sender and subject, followed by a "
            "note for each sender on whether a matching name or company appears "
            "as a SmartSales location."
        ),
        tags=["orchestrator", "multi-agent", "graph", "smartsales"],
    ),
    # graph + salesforce + smartsales - full overview
    Prompt(
        text="Give me an overview of my upcoming meetings from Microsoft 365, related Salesforce opportunities, and SmartSales locations for the companies involved.",
        category="cross-system",
        difficulty="hard",
        expected_answer=(
            "Upcoming calendar events from Microsoft 365, followed by any Salesforce "
            "opportunities related to the meeting subjects or attendees, followed by "
            "SmartSales locations for any recognized company names."
        ),
        tags=["orchestrator", "multi-agent", "graph", "salesforce", "smartsales"],
    ),
]

ORCHESTRATOR_PROMPTS = _ORCHESTRATOR_ROUTING + _ORCHESTRATOR_CROSS


# ── Excel schema ──────────────────────────────────────────────────────────────

EXCEL_FILE = "benchmark_results.xlsx"

# One sheet per agent + summary
AGENT_SHEETS = {
    "GraphAgent":        "Graph",
    "SalesforceAgent":   "Salesforce",
    "SmartSalesAgent":   "SmartSales",
    "OrchestratorAgent": "Orchestrator",
}
SUMMARY_SHEET = "Summary"

RESULT_COLUMNS = [
    "run_id", "timestamp", "prompt",
    "category", "difficulty", "tags",
    "expected_answer",
    "actual_response",
    "response_time_s",
    "input_tokens", "output_tokens", "total_tokens",
    "response_length",
    "llm_score",        # 1–5
    "llm_rationale",
    "llm_comments",
    "success",
    "error",
    "routing_trace",    # JSON string - populated for OrchestratorAgent runs
    "routing_score",    # 1–5 routing correctness (OrchestratorAgent only)
    "routing_rationale",
]

SUMMARY_COLUMNS = [
    "run_id", "timestamp", "agent",
    "prompts_run", "success_rate_%",
    "avg_response_time_s",
    "avg_input_tokens", "avg_output_tokens", "avg_total_tokens",
    "avg_llm_score",
    "avg_routing_score",
]


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _load_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default blank sheet

    agent_sheets = {}
    for agent_key, sheet_name in AGENT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(RESULT_COLUMNS)
            _style_header(ws)
        agent_sheets[agent_key] = wb[sheet_name]

    if SUMMARY_SHEET not in wb.sheetnames:
        ws_s = wb.create_sheet(SUMMARY_SHEET)
        ws_s.append(SUMMARY_COLUMNS)
        _style_header(ws_s)
    summary_sheet = wb[SUMMARY_SHEET]

    return wb, agent_sheets, summary_sheet


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)


# ── LLM evaluator ─────────────────────────────────────────────────────────────

_EVAL_SYSTEM = (
    "You are a benchmark evaluator for an AI agent system. "
    "Your job is to score how well an agent's actual response matches an expected answer."
)

_EVAL_USER_TMPL = """\
Question asked to the agent:
{question}

Expected answer (description of what a correct response should contain):
{expected_answer}

Actual agent response:
{actual_response}

Rate the actual response on a scale of 1 to 5:
  1 – Completely wrong, irrelevant, or no meaningful response
  2 – Partially correct but with major gaps or errors
  3 – Mostly correct with some notable gaps or inaccuracies
  4 – Correct with only minor gaps or formatting differences
  5 – Fully correct and complete

Respond ONLY with valid JSON in this exact format:
{{"score": <integer 1-5>, "rationale": "<one or two sentence justification of the score>", "comments": "<optional broader observations: what was done well, what was missing, or how the response could be improved>"}}
"""


async def evaluate_response(
    openai_client: AsyncAzureOpenAI,
    deployment: str,
    prompt: Prompt,
    actual_response: str,
    success: bool,
) -> tuple[int | None, str, str]:
    """Return (score 1-5, rationale, comments). Returns (None, reason, "") on failure."""
    if not success or not actual_response.strip():
        return 1, "Agent call failed or returned an empty response.", ""

    user_msg = _EVAL_USER_TMPL.format(
        question=prompt.text,
        expected_answer=prompt.expected_answer,
        actual_response=actual_response[:4000],
    )
    try:
        resp = await openai_client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": _EVAL_SYSTEM},
                {"role": "user",   "content": user_msg},
            ],
            temperature=0,
            max_tokens=300,
        )
        raw = resp.choices[0].message.content or ""
        data = json.loads(raw)
        score     = int(data["score"])
        rationale = str(data.get("rationale", ""))
        comments  = str(data.get("comments", ""))
        return score, rationale, comments
    except Exception as exc:
        return None, f"Evaluator error: {exc}", ""


# ── Routing evaluator ─────────────────────────────────────────────────────────

_ROUTING_EVAL_SYSTEM = (
    "You are a benchmark evaluator for a multi-agent AI orchestration system. "
    "Your job is to assess whether an orchestrator correctly routed a user query "
    "to the appropriate sub-agent(s)."
)

_ROUTING_EVAL_USER_TMPL = """\
User query:
{question}

Sub-agents invoked by the orchestrator (in order):
{invoked_agents}

Available sub-agents and their domains:
  graph       – Microsoft 365: emails, OneDrive files, contacts, calendar
  salesforce  – CRM: accounts, contacts, leads, opportunities, cases
  smartsales  – Field sales app: locations, catalog items, orders, approbation statuses

Rate the routing on a scale of 1 to 5:
  1 – Wrong agent(s) called, or no agent called when one was needed
  2 – Partially correct: mostly wrong domain or many unnecessary calls
  3 – Correct agent(s) called but with notable issues (redundant calls, wrong order, missing agent)
  4 – Correct routing with only minor inefficiencies
  5 – Optimal: exactly the right agent(s) called in the right order, no unnecessary calls

Respond ONLY with valid JSON in this exact format:
{{"score": <integer 1-5>, "rationale": "<one or two sentence justification>"}}
"""


def _format_routing_invocations(routing_trace_json: str) -> str:
    try:
        data = json.loads(routing_trace_json)
        invocations = data.get("invoked_agents", [])
        if not invocations:
            return "(none)"
        lines = []
        for inv in invocations:
            status = "success" if inv.get("success") else "FAILED"
            inp = str(inv.get("input", ""))[:200]
            lines.append(f"  {inv['order']}. {inv['agent']} ({status}) - {inp!r}")
        return "\n".join(lines)
    except Exception:
        return "(could not parse trace)"


async def evaluate_routing_response(
    openai_client: AsyncAzureOpenAI,
    deployment: str,
    question: str,
    routing_trace_json: str,
) -> tuple[int | None, str]:
    """Return (routing_score 1-5 or None, rationale)."""
    if not routing_trace_json or not routing_trace_json.strip():
        return None, ""

    try:
        resp = await openai_client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": _ROUTING_EVAL_SYSTEM},
                {"role": "user",   "content": _ROUTING_EVAL_USER_TMPL.format(
                    question=question,
                    invoked_agents=_format_routing_invocations(routing_trace_json),
                )},
            ],
            temperature=0,
            max_tokens=200,
        )
        raw  = resp.choices[0].message.content or ""
        data = json.loads(raw)
        score = data.get("score")
        if score is not None:
            score = int(score)
        return score, str(data.get("rationale", ""))
    except Exception as exc:
        return None, f"Routing evaluator error: {exc}"


# ── Auth / server helpers ─────────────────────────────────────────────────────

_TOKEN_CACHE_FILE = ".token_cache.bin"


def _build_msal_app(client_id: str, tenant_id: str, client_secret: str):
    cache = msal.SerializableTokenCache()
    if os.path.exists(_TOKEN_CACHE_FILE):
        with open(_TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())
    return msal.ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
        token_cache=cache,
    ), cache


def _persist_cache(cache: msal.SerializableTokenCache) -> None:
    if cache.has_state_changed:
        with open(_TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())


def authenticate_microsoft(client_id: str, tenant_id: str, scopes: list[str], client_secret: str) -> str:
    app, cache = _build_msal_app(client_id, tenant_id, client_secret)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
        if result and "access_token" in result:
            _persist_cache(cache)
            return result["access_token"]

    # Auth code flow - open browser, catch redirect on localhost (same as main.py).
    flow = app.initiate_auth_code_flow(scopes=scopes, redirect_uri="http://localhost:5000")
    if "auth_uri" not in flow:
        raise RuntimeError(f"Failed to create auth flow: {flow}")

    print("Opening browser for login...")
    webbrowser.open(flow["auth_uri"])

    auth_response: dict = {}

    class _CallbackHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            params = {k: v[0] for k, v in parse_qs(urlparse(self.path).query).items()}
            auth_response.update(params)
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(b"<html><body><h3>Authenticated! You can close this tab.</h3></body></html>")

        def log_message(self, *args):
            pass

    server = HTTPServer(("localhost", 5000), _CallbackHandler)
    print("Waiting for authentication callback...")
    server.handle_request()
    server.server_close()

    result = app.acquire_token_by_auth_code_flow(flow, auth_response)
    if "access_token" not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description', 'unknown')}")
    _persist_cache(cache)
    return result["access_token"]


def _wait_for_port(host: str, port: int, timeout: float = 15.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        try:
            with socket.create_connection((host, port), timeout=1.0):
                return
        except OSError:
            time.sleep(0.25)
    raise TimeoutError(f"Server at {host}:{port} not ready within {timeout}s")


def _is_local(url: str) -> bool:
    return (urlparse(url).hostname or "") in ("localhost", "127.0.0.1", "::1")


def _start_server(module: str, env: dict, url: str) -> subprocess.Popen:
    proc = subprocess.Popen(
        [sys.executable, "-m", module],
        env=env,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    parsed = urlparse(url)
    _wait_for_port(parsed.hostname or "localhost", parsed.port or 8000)
    return proc


def _resolve_sf_session(sf_mcp_url: str) -> str:
    parsed = urlparse(sf_mcp_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    try:
        resp = httpx.get(f"{base}/auth/salesforce/session", timeout=5)
        if resp.status_code == 200:
            return resp.json()["session_token"]
    except httpx.RequestError as exc:
        raise RuntimeError(f"Cannot reach Salesforce MCP server: {exc}") from exc
    raise RuntimeError(
        "No active Salesforce session. Run main.py first to authenticate via browser."
    )


def _resolve_ss_session(ss_mcp_url: str) -> str:
    parsed = urlparse(ss_mcp_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    try:
        resp = httpx.get(f"{base}/auth/smartsales/session", timeout=10)
        if resp.status_code == 200:
            return resp.json()["session_token"]
        raise RuntimeError(f"SmartSales session returned {resp.status_code}: {resp.text}")
    except httpx.RequestError as exc:
        raise RuntimeError(f"Cannot reach SmartSales MCP server: {exc}") from exc


# ── Core benchmark ────────────────────────────────────────────────────────────

async def run_prompt(agent, prompt: Prompt) -> dict:
    t0 = time.perf_counter()
    response_text = ""
    error = ""
    success = False
    input_tokens = output_tokens = total_tokens = None

    # Initialise a fresh routing trace for every run.  For non-orchestrator
    # agents the trace will be empty (invoked_agents=[]), which is correct.
    trace = start_trace(prompt.text)

    try:
        response = await agent.run(prompt.text)
        response_text = response.text or ""
        print("response_text", response_text)
        success = True

        usage = response.usage_details or {}
        input_tokens  = usage.get("input_token_count")
        output_tokens = usage.get("output_token_count")
        total_tokens  = usage.get("total_token_count")
        if total_tokens is None and input_tokens is not None and output_tokens is not None:
            total_tokens = input_tokens + output_tokens

    except Exception as exc:
        error = str(exc)
        print(f"    ERROR: {exc}")

    # Collect trace after the run (closures have mutated it in-place)
    routing_trace_json = get_trace().to_json() if get_trace() is not None else ""

    return {
        "response_text":   response_text,
        "response_time":   time.perf_counter() - t0,
        "input_tokens":    input_tokens,
        "output_tokens":   output_tokens,
        "total_tokens":    total_tokens,
        "response_length": len(response_text),
        "success":         success,
        "error":           error,
        "routing_trace":   routing_trace_json,
    }


async def benchmark(graph_agent, sf_agent, ss_agent, orchestrator) -> list[dict]:
    timestamp = datetime.now().isoformat(timespec="seconds")
    run_id    = str(uuid.uuid4())[:8]
    results   = []

    modes = [
        ("GraphAgent",        graph_agent,   GRAPH_PROMPTS),
        # ("SalesforceAgent",   sf_agent,      SALESFORCE_PROMPTS),
        # ("SmartSalesAgent",   ss_agent,      SMARTSALES_PROMPTS),
        # ("OrchestratorAgent", orchestrator,  ORCHESTRATOR_PROMPTS),
    ]

    for mode_name, agent, prompts in modes:
        print(f"\n{'-' * 65}")
        print(f"  {mode_name}  ({len(prompts)} prompts)")
        print(f"{'-' * 65}")

        for i, prompt in enumerate(prompts, 1):
            print(f"  [{i:02d}/{len(prompts):02d}] [{prompt.difficulty}] {prompt.text!r}")
            metrics = await run_prompt(agent, prompt)
            status = "OK  " if metrics["success"] else "FAIL"
            print(
                f"           → {status} | {metrics['response_time']:.2f}s "
                f"| tokens in={metrics['input_tokens']} out={metrics['output_tokens']}"
            )
            results.append({
                "run_id":    run_id,
                "timestamp": timestamp,
                "agent_mode": mode_name,
                "prompt":    prompt,   # keep the Prompt object for evaluation step
                **metrics,
            })

    return results


async def evaluate_all(
    results: list[dict],
    openai_client: AsyncAzureOpenAI,
    deployment: str,
) -> None:
    """Add llm_score, llm_rationale, routing_score and routing_rationale to each result dict in-place."""
    total = len(results)
    print(f"\nEvaluating {total} responses with LLM...")
    for i, r in enumerate(results, 1):
        prompt: Prompt = r["prompt"]

        score, rationale, comments = await evaluate_response(
            openai_client, deployment,
            prompt, r["response_text"], r["success"],
        )
        r["llm_score"]     = score
        r["llm_rationale"] = rationale
        r["llm_comments"]  = comments

        routing_trace = r.get("routing_trace", "")
        if routing_trace and routing_trace.strip():
            r_score, r_rationale = await evaluate_routing_response(
                openai_client, deployment,
                prompt.text, routing_trace,
            )
        else:
            r_score, r_rationale = None, ""
        r["routing_score"]     = r_score
        r["routing_rationale"] = r_rationale

        a_status = f"{score}/5"   if score   is not None else "ERR"
        r_status = f"{r_score}/5" if r_score is not None else "-"
        print(f"  [{i:02d}/{total:02d}] answer={a_status}  routing={r_status}  {prompt.text[:55]!r}")


def save_results(results: list[dict]) -> None:
    wb, agent_sheets, summary_sheet = _load_or_create_workbook()

    # Group by agent
    by_agent: dict[str, list[dict]] = {}
    for r in results:
        by_agent.setdefault(r["agent_mode"], []).append(r)

    for agent_mode, rows in by_agent.items():
        ws = agent_sheets[agent_mode]
        for r in rows:
            prompt: Prompt = r["prompt"]
            ws.append([
                r["run_id"],
                r["timestamp"],
                prompt.text,
                prompt.category,
                prompt.difficulty,
                ",".join(prompt.tags),
                prompt.expected_answer,
                r["response_text"],
                round(r["response_time"], 3) if r["response_time"] is not None else None,
                r["input_tokens"],
                r["output_tokens"],
                r["total_tokens"],
                r["response_length"],
                r.get("llm_score"),
                r.get("llm_rationale", ""),
                r.get("llm_comments", ""),
                r["success"],
                r["error"],
                r.get("routing_trace", ""),
                r.get("routing_score"),
                r.get("routing_rationale", ""),
            ])
        _auto_width(ws)

    # Summary
    run_id    = results[0]["run_id"]    if results else ""
    timestamp = results[0]["timestamp"] if results else ""

    for agent_mode, rows in by_agent.items():
        n         = len(rows)
        successes = sum(1 for r in rows if r["success"])
        times  = [r["response_time"]  for r in rows if r["response_time"]  is not None]
        in_t   = [r["input_tokens"]   for r in rows if r["input_tokens"]   is not None]
        out_t  = [r["output_tokens"]  for r in rows if r["output_tokens"]  is not None]
        tot_t  = [r["total_tokens"]   for r in rows if r["total_tokens"]   is not None]
        scores   = [r["llm_score"]       for r in rows if r.get("llm_score")       is not None]
        r_scores = [r["routing_score"]   for r in rows if r.get("routing_score")   is not None]
        summary_sheet.append([
            run_id,
            timestamp,
            agent_mode,
            n,
            round(successes / n * 100, 1) if n else None,
            round(sum(times) / len(times), 3) if times else None,
            round(sum(in_t)  / len(in_t))     if in_t   else None,
            round(sum(out_t) / len(out_t))    if out_t  else None,
            round(sum(tot_t) / len(tot_t))    if tot_t  else None,
            round(sum(scores)   / len(scores),   2) if scores   else None,
            round(sum(r_scores) / len(r_scores), 2) if r_scores else None,
        ])
    _auto_width(summary_sheet)

    target = EXCEL_FILE
    try:
        wb.save(target)
    except PermissionError:
        target = EXCEL_FILE.replace(".xlsx", f"_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb.save(target)
        print(f"\nWARNING: {EXCEL_FILE} is locked. Saved to {target}")
    print(f"\nResults saved → {target}  ({len(results)} rows)")


# ── Routing trace persistence ─────────────────────────────────────────────────

ROUTING_TRACES_FILE = "eval/routing_traces.jsonl"


def save_routing_traces(results: list[dict]) -> None:
    """Append one JSONL line per result that has a non-empty routing_trace.

    Each line is a self-contained JSON object:
      {run_id, timestamp, agent_mode, prompt_text, routing_trace: {...}}

    This file is easy to consume from any analysis script:
      import json
      traces = [json.loads(l) for l in open("eval/routing_traces.jsonl")]
    """
    lines = []
    for r in results:
        raw = r.get("routing_trace", "")
        if not raw:
            continue
        try:
            trace_dict = json.loads(raw)
        except json.JSONDecodeError:
            continue
        prompt: Prompt = r["prompt"]
        lines.append(json.dumps({
            "run_id":      r["run_id"],
            "timestamp":   r["timestamp"],
            "agent_mode":  r["agent_mode"],
            "prompt_text": prompt.text,
            "routing_trace": trace_dict,
        }, ensure_ascii=False))

    if not lines:
        return

    os.makedirs(os.path.dirname(ROUTING_TRACES_FILE), exist_ok=True)
    with open(ROUTING_TRACES_FILE, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    print(f"Routing traces appended → {ROUTING_TRACES_FILE}  ({len(lines)} entries)")


# ── Entry point ───────────────────────────────────────────────────────────────

async def main() -> None:
    config = configparser.ConfigParser()
    config.read(["config.cfg"])
    azure  = config["azure"]
    sf_cfg = config["salesforce"]
    ss_cfg = config["smartsales"] if config.has_section("smartsales") else {}

    deployment = os.environ["deployment"]
    endpoint   = os.environ["AZURE_OPENAI_ENDPOINT"]
    api_key    = os.environ["AZURE_OPENAI_API_KEY"]

    total = (
        len(GRAPH_PROMPTS) + len(SALESFORCE_PROMPTS)
        + len(SMARTSALES_PROMPTS) + len(ORCHESTRATOR_PROMPTS)
    )
    print(
        f"\nBenchmark plan: {len(GRAPH_PROMPTS)} graph + {len(SALESFORCE_PROMPTS)} salesforce"
        f" + {len(SMARTSALES_PROMPTS)} smartsales + {len(ORCHESTRATOR_PROMPTS)} orchestrator"
        f" = {total} prompts\n"
    )

    # ── Microsoft Graph ────────────────────────────────────────────────────────
    client_id     = azure["clientId"]
    tenant_id     = azure["tenantId"]
    client_secret = azure.get("clientSecret", os.environ.get("CLIENT_SECRET", ""))
    scopes        = azure["graphUserScopes"].split()
    graph_mcp_url = azure.get("mcpServerUrl", "http://localhost:8000/mcp")

    print("Authenticating with Microsoft...")
    ms_token = authenticate_microsoft(client_id, tenant_id, scopes, client_secret)
    print("OK")

    graph_env = os.environ.copy()
    parsed    = urlparse(graph_mcp_url)
    graph_env["MCP_RESOURCE_URI"] = f"{parsed.scheme}://{parsed.netloc}"

    graph_proc = None
    if _is_local(graph_mcp_url):
        print("Starting Graph MCP server...")
        graph_proc = _start_server("graph.mcp_server", graph_env, graph_mcp_url)
        print("OK")

    graph_http = httpx.AsyncClient(headers={"Authorization": f"Bearer {ms_token}"})
    graph_mcp  = MCPStreamableHTTPTool(name="graph", url=graph_mcp_url, http_client=graph_http)

    # ── Salesforce ─────────────────────────────────────────────────────────────
    sf_mcp_url = sf_cfg.get("mcpServerUrl", "http://localhost:8001/mcp")
    sf_env     = os.environ.copy()
    sf_parsed  = urlparse(sf_mcp_url)
    sf_env["MCP_RESOURCE_URI"] = f"{sf_parsed.scheme}://{sf_parsed.netloc}"

    sf_proc = None
    if _is_local(sf_mcp_url):
        print("Starting Salesforce MCP server...")
        sf_proc = _start_server("salesforce.mcp_server", sf_env, sf_mcp_url)
        print("OK")

    print("Resolving Salesforce session...")
    sf_token = _resolve_sf_session(sf_mcp_url)
    print("OK")

    sf_http = httpx.AsyncClient(headers={"Authorization": f"Bearer {sf_token}"})
    sf_mcp  = MCPStreamableHTTPTool(name="salesforce", url=sf_mcp_url, http_client=sf_http)

    # ── SmartSales ─────────────────────────────────────────────────────────────
    ss_mcp_url = ss_cfg.get("mcpServerUrl", "http://localhost:8002/mcp")
    ss_env     = os.environ.copy()
    ss_parsed  = urlparse(ss_mcp_url)
    ss_env["MCP_RESOURCE_URI"] = f"{ss_parsed.scheme}://{ss_parsed.netloc}"

    ss_proc = None
    if _is_local(ss_mcp_url):
        print("Starting SmartSales MCP server...")
        ss_proc = _start_server("smartsales.mcp_server", ss_env, ss_mcp_url)
        print("OK")

    print("Resolving SmartSales session...")
    ss_token = _resolve_ss_session(ss_mcp_url)
    print("OK")

    ss_http = httpx.AsyncClient(headers={"Authorization": f"Bearer {ss_token}"})
    ss_mcp  = MCPStreamableHTTPTool(name="smartsales", url=ss_mcp_url, http_client=ss_http)

    # ── Build agents ───────────────────────────────────────────────────────────
    graph_agent  = create_graph_agent(graph_mcp=graph_mcp)
    sf_agent     = create_salesforce_agent(salesforce_mcp=sf_mcp)
    ss_agent     = create_smartsales_agent(smartsales_mcp=ss_mcp)
    orchestrator = create_orchestrator_agent(
        graph_agent=graph_agent,
        salesforce_agent=sf_agent,
        smartsales_agent=ss_agent,
    )

    # ── LLM evaluator client ───────────────────────────────────────────────────
    eval_client = AsyncAzureOpenAI(
        azure_endpoint=endpoint,
        api_key=api_key,
        api_version="2024-12-01-preview",
    )

    # ── Run ────────────────────────────────────────────────────────────────────
    try:
        print(f"Starting benchmark - {datetime.now():%Y-%m-%d %H:%M:%S}")
        results = await benchmark(graph_agent, sf_agent, ss_agent, orchestrator)
        await evaluate_all(results, eval_client, deployment)
        save_results(results)
        save_routing_traces(results)
    finally:
        await graph_http.aclose()
        await sf_http.aclose()
        await ss_http.aclose()
        for proc in (graph_proc, sf_proc, ss_proc):
            if proc is not None:
                proc.terminate()
                proc.wait()


def _exception_handler(loop, context):
    if context.get("message") == "an error occurred during closing of asynchronous generator":
        asyncgen = context.get("asyncgen")
        filename = getattr(getattr(asyncgen, "ag_code", None), "co_filename", "")
        if "streamable_http" in filename:
            return
    loop.default_exception_handler(context)


if __name__ == "__main__":
    loop = asyncio.new_event_loop()
    loop.set_exception_handler(_exception_handler)
    try:
        loop.run_until_complete(main())
    finally:
        loop.close()
