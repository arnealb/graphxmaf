"""eval/score.py — Standalone LLM evaluator for benchmark_results.xlsx.

Reads agent responses that were collected by eval/script.py and fills in
llm_score (1-5) and llm_rationale for any row that has not been scored yet.

Usage (from project root):
    python eval/score.py                     # score only unscored rows
    python eval/score.py --force             # re-score every row
    python eval/score.py --run-id abc12345   # only score a specific run
    python eval/score.py --sheet SmartSales  # only score one agent sheet
"""

import argparse
import asyncio
import json
import os

import openpyxl
from dotenv import load_dotenv
from openai import AsyncAzureOpenAI

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_FILE = "benchmark_results.xlsx"

AGENT_SHEETS = ["Graph", "Salesforce", "SmartSales", "Orchestrator"]

# Column positions are read dynamically from the header row — no hardcoding needed.

# ── Evaluator prompt ──────────────────────────────────────────────────────────

_SYSTEM = (
    "You are a benchmark evaluator for a multi-agent AI system. "
    "Your job is to score how well an agent's actual response matches an expected answer. "
    "Important: if the response states 'no results found' or 'no recent email found' for a system, "
    "that counts as the system being queried — do NOT treat it as missing coverage. "
    "Only penalise for missing coverage if the system was clearly never consulted at all."
)

_USER_TMPL = """\
Question asked to the agent:
{question}

Expected answer (description of what a correct response should contain):
{expected_answer}

Actual agent response:
{actual_response}

Rate the actual response on a scale of 1 to 5:
  1 – Completely wrong, irrelevant, or no meaningful response
  2 – Partially correct but with major gaps or errors
  3 – Mostly correct with some notable gaps or inaccuracies
  4 – Correct with only minor gaps or formatting differences
  5 – Fully correct and complete

Respond ONLY with valid JSON in this exact format:
{{"score": <integer 1-5>, "rationale": "<one or two sentence justification of the score>", "comments": "<optional broader observations: what was done well, what was missing, or how the response could be improved>"}}
"""


async def evaluate(
    client: AsyncAzureOpenAI,
    deployment: str,
    question: str,
    expected_answer: str,
    actual_response: str,
    success: bool,
) -> tuple[int | None, str, str]:
    """Return (score 1-5, rationale, comments)."""
    if not success or not (actual_response or "").strip():
        return 1, "Agent call failed or returned an empty response.", ""

    try:
        resp = await client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": _SYSTEM},
                {"role": "user",   "content": _USER_TMPL.format(
                    question=question,
                    expected_answer=expected_answer,
                    actual_response=actual_response[:4000],
                )},
            ],
            temperature=0,
            max_tokens=300,
        )
        raw  = resp.choices[0].message.content or ""
        data = json.loads(raw)
        return int(data["score"]), str(data.get("rationale", "")), str(data.get("comments", ""))
    except Exception as exc:
        return None, f"Evaluator error: {exc}", ""


# ── Routing evaluator ─────────────────────────────────────────────────────────

_ROUTING_SYSTEM = (
    "You are a benchmark evaluator for a multi-agent AI orchestration system. "
    "Your job is to assess whether an orchestrator correctly routed a user query "
    "to the appropriate sub-agent(s)."
)

_ROUTING_USER_TMPL = """\
User query:
{question}

Sub-agents invoked by the orchestrator (in order):
{invoked_agents}

Expected sub-agents (ground truth):
{expected_agents}

Available sub-agents and their domains:
  graph       – Microsoft 365: emails, OneDrive files, contacts, calendar
  salesforce  – CRM: accounts, contacts, leads, opportunities, cases
  smartsales  – Field sales app: locations, catalog items, orders, approbation statuses

Rate the routing on a scale of 1 to 5 based on how well the invoked agents match the expected agents:
  1 – Wrong agent(s) called, or a required agent was completely missed
  2 – Partially correct: at least one required agent missing or many unnecessary calls
  3 – Correct agent(s) called but with notable issues (redundant calls, wrong order, missing agent)
  4 – Correct routing with only minor inefficiencies
  5 – Optimal: exactly the expected agent(s) called, no unnecessary calls

Respond ONLY with valid JSON in this exact format:
{{"score": <integer 1-5>, "rationale": "<one or two sentence justification>"}}
"""


def _format_invocations(routing_trace_json: str) -> str:
    """Format a routing trace JSON string for the evaluator prompt."""
    try:
        data = json.loads(routing_trace_json)
        invocations = data.get("invoked_agents", [])
        if not invocations:
            return "(none)"
        lines = []
        for inv in invocations:
            status = "success" if inv.get("success") else "FAILED"
            inp = str(inv.get("input", ""))[:200]
            lines.append(f"  {inv['order']}. {inv['agent']} ({status}) — {inp!r}")
        return "\n".join(lines)
    except Exception:
        return "(could not parse trace)"


async def evaluate_routing(
    client: AsyncAzureOpenAI,
    deployment: str,
    question: str,
    routing_trace_json: str,
    expected_agents: list[str] | None = None,
) -> tuple[int | None, str]:
    """Return (routing_score 1-5 or None, rationale)."""
    if not routing_trace_json or not routing_trace_json.strip():
        return None, "No routing trace available."

    invoked_agents_str = _format_invocations(routing_trace_json)
    if expected_agents:
        expected_agents_str = ", ".join(expected_agents)
    else:
        expected_agents_str = "(not specified — infer from the query)"
    try:
        resp = await client.chat.completions.create(
            model=deployment,
            messages=[
                {"role": "system", "content": _ROUTING_SYSTEM},
                {"role": "user",   "content": _ROUTING_USER_TMPL.format(
                    question=question,
                    invoked_agents=invoked_agents_str,
                    expected_agents=expected_agents_str,
                )},
            ],
            temperature=0,
            max_tokens=200,
        )
        raw  = resp.choices[0].message.content or ""
        data = json.loads(raw)
        score = data.get("score")
        if score is not None:
            score = int(score)
        return score, str(data.get("rationale", ""))
    except Exception as exc:
        return None, f"Routing evaluator error: {exc}"


# ── Sheet processing ──────────────────────────────────────────────────────────

def _build_col_map(ws) -> dict[str, int]:
    """Read column positions from the header row (row 1). Returns {name: 0-based index}."""
    return {
        cell.value: cell.column - 1
        for cell in ws[1]
        if cell.value is not None
    }


def _cell(row, col_name: str, col_map: dict) -> object:
    """Read a cell value by column name using the dynamic column map."""
    idx = col_map.get(col_name)
    return row[idx].value if idx is not None and idx < len(row) else None


def _is_scored(row, col_map: dict) -> bool:
    score = _cell(row, "llm_score", col_map)
    return score is not None and str(score).strip() != ""


def _is_routing_scored(row, col_map: dict) -> bool:
    score = _cell(row, "routing_score", col_map)
    return score is not None and str(score).strip() != ""


async def score_sheet(
    ws,
    client: AsyncAzureOpenAI,
    deployment: str,
    force: bool,
    run_id_filter: str | None,
    skip_routing: bool = False,
) -> tuple[int, int]:
    """Score all eligible rows in a worksheet. Returns (scored, skipped)."""
    col_map = _build_col_map(ws)
    scored  = 0
    skipped = 0

    # Collect rows that need any scoring (answer quality and/or routing)
    rows_to_process: list[tuple] = []
    for row in ws.iter_rows(min_row=2):
        run_id = _cell(row, "run_id", col_map)

        if run_id_filter and str(run_id) != run_id_filter:
            skipped += 1
            continue

        needs_answer  = force or not _is_scored(row, col_map)
        routing_trace = str(_cell(row, "routing_trace", col_map) or "")
        needs_routing = (
            not skip_routing
            and bool(routing_trace.strip())
            and (force or not _is_routing_scored(row, col_map))
        )

        if not needs_answer and not needs_routing:
            skipped += 1
            continue

        rows_to_process.append((row, needs_answer, needs_routing))

    total = len(rows_to_process)
    for i, (row, needs_answer, needs_routing) in enumerate(rows_to_process, 1):
        question      = str(_cell(row, "prompt",    col_map) or "")
        success_val   = _cell(row, "success",       col_map)
        success       = bool(success_val) if success_val is not None else False
        run_id        = _cell(row, "run_id",        col_map)
        difficulty    = _cell(row, "difficulty",    col_map)
        excel_row     = row[0].row

        print(f"  [{i:02d}/{total:02d}] run={run_id}  [{difficulty}]  {question[:60]!r}")

        updates: dict = {}

        if needs_answer:
            expected_answer = str(_cell(row, "expected_answer", col_map) or "")
            actual_response = str(_cell(row, "actual_response", col_map) or "")
            score, rationale, comments = await evaluate(
                client, deployment,
                question, expected_answer, actual_response, success,
            )
            updates["llm_score"]     = score
            updates["llm_rationale"] = rationale
            updates["llm_comments"]  = comments
            label = f"{score}/5" if score is not None else "ERR"
            print(f"           answer  → {label}  {rationale[:70]}")

        if needs_routing:
            routing_trace = str(_cell(row, "routing_trace", col_map) or "")
            raw_expected = _cell(row, "expected_agents", col_map)
            if raw_expected:
                exp_agents = [a.strip() for a in str(raw_expected).split(",") if a.strip()]
            else:
                exp_agents = None
            r_score, r_rationale = await evaluate_routing(
                client, deployment,
                question, routing_trace,
                expected_agents=exp_agents,
            )
            updates["routing_score"]     = r_score
            updates["routing_rationale"] = r_rationale
            r_label = f"{r_score}/5" if r_score is not None else "ERR"
            print(f"           routing → {r_label}  {r_rationale[:70]}")

        # Write scores back — add missing columns on the fly for backward compatibility
        for col_name, value in updates.items():
            if col_name not in col_map:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col).value = col_name
                col_map[col_name] = new_col - 1
            ws.cell(row=excel_row, column=col_map[col_name] + 1).value = value

        scored += 1

    return scored, skipped


# ── Entry point ───────────────────────────────────────────────────────────────

async def main(args: argparse.Namespace) -> None:
    skip_routing: bool = getattr(args, "no_routing", False)
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: {EXCEL_FILE} not found. Run eval/script.py first.")
        return

    deployment = os.environ["deployment"]
    endpoint   = os.environ["AZURE_OPENAI_ENDPOINT"]
    api_key    = os.environ["AZURE_OPENAI_API_KEY"]

    client = AsyncAzureOpenAI(
        azure_endpoint=endpoint,
        api_key=api_key,
        api_version="2024-12-01-preview",
    )

    wb     = openpyxl.load_workbook(EXCEL_FILE)
    sheets = [args.sheet] if args.sheet else AGENT_SHEETS

    total_scored  = 0
    total_skipped = 0

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found — skipping.")
            continue

        ws = wb[sheet_name]
        col_map = _build_col_map(ws)
        unscored = sum(
            1 for row in ws.iter_rows(min_row=2)
            if not _is_scored(row, col_map)
            and (not args.run_id or str(_cell(row, "run_id", col_map)) == args.run_id)
        )

        if not args.force and unscored == 0:
            print(f"\n[{sheet_name}] — all rows already scored, skipping. (use --force to re-score)")
            continue

        mode = "re-scoring all" if args.force else f"{unscored} unscored"
        print(f"\n{'─' * 60}")
        print(f"  Sheet: {sheet_name}  ({mode} rows)")
        print(f"{'─' * 60}")

        scored, skipped = await score_sheet(
            ws, client, deployment,
            force=args.force,
            run_id_filter=args.run_id,
            skip_routing=skip_routing,
        )
        total_scored  += scored
        total_skipped += skipped

    # Save
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        from datetime import datetime
        fallback = EXCEL_FILE.replace(".xlsx", f"_scored_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb.save(fallback)
        print(f"\nWARNING: {EXCEL_FILE} is locked. Saved to {fallback}")
        return

    print(f"\nDone — scored {total_scored} rows, skipped {total_skipped}.")
    print(f"Results saved → {EXCEL_FILE}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Score benchmark_results.xlsx with LLM evaluator.")
    parser.add_argument(
        "--force", action="store_true",
        help="Re-score rows that already have a score.",
    )
    parser.add_argument(
        "--run-id", metavar="ID",
        help="Only score rows matching this run_id (e.g. abc12345).",
    )
    parser.add_argument(
        "--sheet", choices=["Graph", "Salesforce", "SmartSales", "Orchestrator"],
        help="Only score rows in this agent sheet.",
    )
    parser.add_argument(
        "--no-routing", action="store_true",
        help="Skip routing correctness scoring (only score answer quality).",
    )
    args = parser.parse_args()
    asyncio.run(main(args))
