"""
evaluate_agents.py — Benchmark GraphAgent, SalesforceAgent, and OrchestratorAgent.

Each run appends rows to benchmark_results.xlsx so results accumulate across
multiple invocations and can be compared over time.

Usage:
    python evaluate_agents.py
"""

import asyncio
import configparser
import os
import socket
import subprocess
import sys
import time
import uuid
from datetime import datetime
from urllib.parse import urlparse

import httpx
import msal
import openpyxl
from dotenv import load_dotenv

from agent_framework import MCPStreamableHTTPTool
from agents.graph_agent import create_graph_agent
from agents.orchestrator_agent import create_orchestrator_agent
from agents.salesforce_agent import create_salesforce_agent
from salesforce.auth import authenticate_salesforce

load_dotenv()

# ── Prompts ────────────────────────────────────────────────────────────────────

GRAPH_PROMPTS = [
    "Who am I in Microsoft 365?",
    "List my 5 most recent emails",
    "Find the email address of dorian",
    "What are my calendar events for next week?"
]

SALESFORCE_PROMPTS = [
    "Fetch 5 Salesforce accounts in the technology sector",
    "give me 5 salesforce opportunities with an amount greater than 10k",
]

# The orchestrator is tested on all prompts so routing behaviour is exercised.
ALL_PROMPTS = GRAPH_PROMPTS + SALESFORCE_PROMPTS

EXCEL_FILE = "benchmark_results.xlsx"
SHEET_NAME = "Results"
COLUMNS = [
    "run_id",
    "timestamp",
    "agent_mode",
    "prompt",
    "response_time_s",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "success",
    "error",
    "response",
]

# ── Excel helpers ──────────────────────────────────────────────────────────────

def _load_or_create_workbook() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)
        _style_header(ws)
    return wb, ws


def _style_header(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _append_row(ws, run_id: str, timestamp: str, agent_mode: str, prompt: str,
                response_time: float | None, input_tokens: int | None,
                output_tokens: int | None, total_tokens: int | None,
                success: bool, error: str, response_text: str) -> None:
    ws.append([
        run_id,
        timestamp,
        agent_mode,
        prompt,
        round(response_time, 3) if response_time is not None else None,
        input_tokens,
        output_tokens,
        total_tokens,
        success,
        error,
        response_text,
    ])


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

# ── Auth / MCP helpers (mirrors main.py) ──────────────────────────────────────

_TOKEN_CACHE_FILE = ".token_cache.bin"


def _build_msal_app(client_id: str, tenant_id: str):
    cache = msal.SerializableTokenCache()
    if os.path.exists(_TOKEN_CACHE_FILE):
        with open(_TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())
    app = msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        token_cache=cache,
    )
    return app, cache


def _persist_cache(cache: msal.SerializableTokenCache) -> None:
    if cache.has_state_changed:
        with open(_TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())


def authenticate_microsoft(client_id: str, tenant_id: str, scopes: list[str]) -> str:
    app, cache = _build_msal_app(client_id, tenant_id)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
        if result and "access_token" in result:
            _persist_cache(cache)
            return result["access_token"]
    flow = app.initiate_device_flow(scopes=scopes)
    print(f"\nAuthenticate at: {flow['verification_uri']}")
    print(f"Enter code:      {flow['user_code']}\n")
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description', 'unknown')}")
    _persist_cache(cache)
    return result["access_token"]


def _wait_for_port(host: str, port: int, timeout: float = 15.0) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        try:
            with socket.create_connection((host, port), timeout=1.0):
                return
        except OSError:
            time.sleep(0.25)
    raise TimeoutError(f"MCP server at {host}:{port} did not become ready within {timeout}s")


def _is_local(url: str) -> bool:
    host = urlparse(url).hostname or ""
    return host in ("localhost", "127.0.0.1", "::1")


def _start_server(module: str, env: dict, url: str) -> subprocess.Popen:
    proc = subprocess.Popen(
        [sys.executable, "-m", module],
        env=env,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    parsed = urlparse(url)
    _wait_for_port(parsed.hostname or "localhost", parsed.port or 8000)
    return proc

# ── Core benchmark logic ───────────────────────────────────────────────────────

async def run_prompt(agent, prompt: str) -> dict:
    """Send one prompt to an agent and return timing + token metrics."""
    t0 = time.perf_counter()
    error = ""
    response_time = None
    input_tokens = output_tokens = total_tokens = None
    success = False

    try:
        response = await agent.run(prompt)
        response_time = time.perf_counter() - t0
        success = True

        usage = response.usage_details or {}
        input_tokens = usage.get("input_token_count")
        output_tokens = usage.get("output_token_count")
        total_tokens = usage.get("total_token_count")

        # Fall back to summing if total is missing
        if total_tokens is None and input_tokens is not None and output_tokens is not None:
            total_tokens = input_tokens + output_tokens

        response_text = response.text or ""

    except Exception as exc:
        response_time = time.perf_counter() - t0
        error = str(exc)
        print(f"    ERROR: {exc}")

    return {
        "response_text": response_text,
        "response_time": response_time,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
        "success": success,
        "error": error,
    }


async def benchmark(graph_agent, sf_agent, orchestrator_agent) -> list[dict]:
    """Run all prompts against all agent modes and collect results."""
    timestamp = datetime.now().isoformat(timespec="seconds")
    run_id = str(uuid.uuid4())[:8]
    results = []

    modes = [
        ("GraphAgent",       graph_agent,        GRAPH_PROMPTS),
        ("SalesforceAgent",  sf_agent,            SALESFORCE_PROMPTS),
        ("OrchestratorAgent", orchestrator_agent, ALL_PROMPTS),
    ]

    for mode_name, agent, prompts in modes:
        print(f"\n{'─' * 60}")
        print(f"  Mode: {mode_name}  ({len(prompts)} prompts)")
        print(f"{'─' * 60}")

        for i, prompt in enumerate(prompts, 1):
            print(f"  [{i}/{len(prompts)}] {prompt!r}")
            metrics = await run_prompt(agent, prompt)
            status = "OK" if metrics["success"] else "FAIL"
            print(
                f"         → {status} | {metrics['response_time']:.2f}s "
                f"| tokens: in={metrics['input_tokens']} "
                f"out={metrics['output_tokens']} "
                f"total={metrics['total_tokens']}"
            )
            results.append({
                "run_id": run_id,
                "timestamp": timestamp,
                "agent_mode": mode_name,
                "prompt": prompt,
                **metrics,
            })

    return results


def save_results(results: list[dict]) -> None:
    wb, ws = _load_or_create_workbook()
    for r in results:
        _append_row(
            ws,
            run_id=r["run_id"],
            timestamp=r["timestamp"],
            agent_mode=r["agent_mode"],
            prompt=r["prompt"],
            response_time=r["response_time"],
            input_tokens=r["input_tokens"],
            output_tokens=r["output_tokens"],
            total_tokens=r["total_tokens"],
            success=r["success"],
            error=r["error"],
            response_text=r.get("response_text", ""),
        )
    _auto_width(ws)
    target = EXCEL_FILE
    try:
        wb.save(target)
    except PermissionError:
        # File is likely open in Excel — save to a timestamped fallback.
        target = EXCEL_FILE.replace(".xlsx", f"_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb.save(target)
        print(f"\nWARNING: {EXCEL_FILE} is locked (open in Excel?). Saved to fallback file.")
    print(f"Results saved → {target}  ({len(results)} new rows)")

# ── Entry point ────────────────────────────────────────────────────────────────

async def main() -> None:
    config = configparser.ConfigParser()
    config.read(["config.cfg"])
    azure = config["azure"]
    sf_cfg = config["salesforce"]

    # ── Microsoft Graph auth & MCP server ─────────────────────────────────────
    client_id = azure["clientId"]
    tenant_id = azure["tenantId"]
    scopes = azure["graphUserScopes"].split()
    graph_mcp_url = azure.get("mcpServerUrl", "http://localhost:8000/mcp")

    print("Authenticating with Microsoft…")
    ms_token = authenticate_microsoft(client_id, tenant_id, scopes)
    print("OK")

    server_env = os.environ.copy()
    parsed = urlparse(graph_mcp_url)
    server_env["MCP_RESOURCE_URI"] = f"{parsed.scheme}://{parsed.netloc}"

    graph_proc = None
    if _is_local(graph_mcp_url):
        print("Starting Graph MCP server…")
        graph_proc = _start_server("graph.mcp_server", server_env, graph_mcp_url)
        print("OK")

    graph_http = httpx.AsyncClient(headers={"Authorization": f"Bearer {ms_token}"})
    graph_mcp = MCPStreamableHTTPTool(name="graph", url=graph_mcp_url, http_client=graph_http)

    # ── Salesforce auth & MCP server ──────────────────────────────────────────
    sf_mcp_url = sf_cfg.get("mcpServerUrl", "http://localhost:8001/mcp")
    sf_login_url = sf_cfg.get("loginUrl", "https://test.salesforce.com")

    print("Authenticating with Salesforce…")
    sf_creds = authenticate_salesforce(login_url=sf_login_url)
    print("OK")

    sf_env = os.environ.copy()
    sf_parsed = urlparse(sf_mcp_url)
    sf_env["MCP_RESOURCE_URI"] = f"{sf_parsed.scheme}://{sf_parsed.netloc}"
    sf_env["SF_INSTANCE_URL"] = sf_creds.instance_url

    sf_proc = None
    if _is_local(sf_mcp_url):
        print("Starting Salesforce MCP server…")
        sf_proc = _start_server("salesforce.mcp_server", sf_env, sf_mcp_url)
        print("OK")

    sf_http = httpx.AsyncClient(headers={"Authorization": f"Bearer {sf_creds.access_token}"})
    sf_mcp = MCPStreamableHTTPTool(name="salesforce", url=sf_mcp_url, http_client=sf_http)

    # ── Build agents ──────────────────────────────────────────────────────────
    graph_agent = create_graph_agent(graph_mcp=graph_mcp)
    sf_agent = create_salesforce_agent(salesforce_mcp=sf_mcp)
    orchestrator = create_orchestrator_agent(graph_agent=graph_agent, salesforce_agent=sf_agent)

    # ── Run benchmark ─────────────────────────────────────────────────────────
    try:
        print(f"\nStarting benchmark — {datetime.now():%Y-%m-%d %H:%M:%S}")
        results = await benchmark(graph_agent, sf_agent, orchestrator)
        save_results(results)


    # finally:
    #     await graph_http.aclose()
    #     await sf_http.aclose()
    #     for proc in (graph_proc, sf_proc):
    #         if proc is not None:
    #             proc.terminate()
    #             proc.wait()

    finally:
        # Close HTTP clients; MCP tool cleanup is handled by server termination below.
        await graph_http.aclose()
        await sf_http.aclose()

        for proc in (graph_proc, sf_proc):
            if proc is not None:
                proc.terminate()
                proc.wait()


def _exception_handler(loop, context):
    # Suppress the anyio cancel-scope errors that occur when the MCP
    # streamable-HTTP async generators are garbage-collected after the
    # event loop closes.  Everything else goes to the default handler.
    if context.get("message") == "an error occurred during closing of asynchronous generator":
        asyncgen = context.get("asyncgen")
        filename = getattr(getattr(asyncgen, "ag_code", None), "co_filename", "")
        if "streamable_http" in filename:
            return
    loop.default_exception_handler(context)


if __name__ == "__main__":
    loop = asyncio.new_event_loop()
    loop.set_exception_handler(_exception_handler)
    try:
        loop.run_until_complete(main())
    finally:
        loop.close()
